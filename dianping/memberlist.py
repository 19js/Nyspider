import requests
from bs4 import BeautifulSoup
import time
from proxy import *
import datetime
import openpyxl

headers = {
    'Host':'www.dianping.com',
    'User-Agent':"Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:44.0) Gecko/20100101 Firefox/44.0",
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.5',
    'Accept-Encoding': 'gzip, deflate',
    'Connection': 'keep-alive'}

def get_memberlist():
    page=1
    while True:
        url='http://www.dianping.com/memberlist/4/10?pg='+str(page)
        html=requests.get(url,headers=headers).text
        table=BeautifulSoup(html,'lxml').find('table',{'class':'rankTable'}).find('tbody').find_all('tr')
        f=open('memberlist.txt','a')
        for item in table:
            try:
                tds=item.find_all('td')
                name=tds[0].find('a').get_text()
                url=tds[0].find('a').get('href')
                comment_num=tds[1].get_text()
                reply_num=tds[3].get_text()
                flower_num=tds[4].get_text()
                f.write(str([name,url,comment_num,reply_num,flower_num])+'\n')
            except:
                continue
        f.close()
        page+=1
        if page==7:
            break

def get_comments(usrid):
    baseurl='http://www.dianping.com/member/{}/reviews?pg={}&reviewCityId=0&reviewShopType=10&c=0&shopTypeIndex=1'
    page=1
    html=requests.get(baseurl.format(usrid,page),headers=headers,proxies=get_proxies()).text
    soup=BeautifulSoup(html,'lxml').find('div',{'class':'main'})
    citys=soup.find('div',{'class':'p-term-list'}).find_all('li')[1].find_all('span')
    city_num=len(citys)
    city_text=''
    for span in citys:
        city_text+=span.get_text()+'\t'
    table=soup.find('div',id='J_review').find_all('div',{'class':'J_rptlist'})
    result=[]
    for item in table:
        title=item.find('a').get_text()
        url=item.find('a').get('href')
        try:
            address=item.find('div',{'class':'addres'}).get_text()
        except:
            address=''
        try:
            star=item.find('div',{'class':'comm-rst'}).find('span').get('class')[1].replace('irr-star','')
        except:
            star=''
        content=item.find('div',{'class':'comm-entry'}).get_text()
        date=item.find('div',{'class':"info"}).find('span').get_text().replace("发表于",'')
        line=[city_num,city_text,title,url,address,star,content,date]
        result.append(line)
    page+=1
    while True:
        print(page)
        html=requests.get(baseurl.format(usrid,page),headers=headers,proxies=get_proxies()).text
        soup=BeautifulSoup(html,'lxml').find('div',{'class':'main'})
        table=soup.find('div',id='J_review').find_all('div',{'class':'J_rptlist'})
        for item in table:
            title=item.find('a').get_text()
            url=item.find('a').get('href')
            try:
                address=item.find('div',{'class':'addres'}).get_text()
            except:
                address=''
            try:
                star=item.find('div',{'class':'comm-rst'}).find('span').get('class')[1].replace('irr-star','')
            except:
                star=''
            content=item.find('div',{'class':'comm-entry'}).get_text()
            date=item.find('div',{'class':"info"}).find('span').get_text().replace("发表于",'')
            line=[city_num,city_text,title,url,address,star,content,date]
            result.append(line)
            if len(result)==100:
                return result
        page+=1
    return result

def shop_infor(shopurl):
    while True:
        try:
            html=requests.get(shopurl,headers=headers,proxies=get_proxies(),timeout=30).text
            if '您使用的IP访问网站过于频繁，为了您的正常访问，请先输入验证码' in html:
                switch_ip()
                continue
            break
        except:
            switch_ip()
            continue
    soup=BeautifulSoup(html,'lxml').find('div',{'class':'body-content'})
    try:
        types=soup.find('div',{'class':'breadcrumb'}).find_all('a')
        shop_type=types[2].get_text()
    except:
        shop_type=''
    base_infor=soup.find('div',id='basic-info').find('div',{'class':'brief-info'})
    try:
        star=base_infor.find('span',{'class':'mid-rank-stars'}).get('class')[1].replace('mid-str','')
    except:
        star=''
    line=['']*4
    for item in base_infor.find_all('span'):
        if '人均' in str(item):
            text=item.get_text()
            line[0]=text
        if '口味' in str(item):
            text=item.get_text()
            line[1]=text
        if '环境' in str(item):
            text=item.get_text()
            line[2]=text
        if '服务' in str(item):
            text=item.get_text()
            line[3]=text
    return [shop_type,star]+line

def get_fans(usrid):
    url='http://www.dianping.com/member/{}/fans?pg={}'
    page=1
    fans=[]
    while True:
        try:
            html=requests.get(url.format(usrid,page),headers=headers,timeout=30,proxies=get_proxies()).text
            if '您使用的IP访问网站过于频繁，为了您的正常访问，请先输入验证码' in html:
                switch_ip()
                continue
        except:
            continue
        try:
            table=BeautifulSoup(html,'lxml').find('div',{'class':'main'}).find('div',{'class':'pic-txt'}).find_all("li")
        except:
            break
        for item in table:
            try:
                name=item.find('h6').get_text()
                fans.append(name)
            except:
                continue
        page+=1
        print(usrid,page)
    return [len(fans),fans]

def get_week_day(date_str):
    date=datetime.datetime.strptime(date_str,'%Y-%m-%d')
    week_day_dict={
        0 : '星期一',
        1 : '星期二',
        2 : '星期三',
        3 : '星期四',
        4 : '星期五',
        5 : '星期六',
        6 : '星期天',
    }
    day = date.weekday()
    return week_day_dict[day]

def main():
    '''
    usrs=[eval(line) for line in open('./memberlist.txt','r')]
    for usr in usrs:
        usrid=usr[1].split('/')[-1]
        try:
            result=get_comments(usrid)
        except:
            failed=open('failed.txt','a')
            failed.write(str(usr)+'\n')
            failed.close()
            print(usr,'failed')
            continue
        f=open('comments.txt','a')
        for item in result:
            f.write(str(usr+item)+'\n')
        f.close()
        switch_ip()
        print(usr,'ok')
    for line in open('./shop_failed.txt','r'):
        item=eval(line)
        date_str='20'+item[-1]
        try:
            weekday=get_week_day(date_str)
        except:
            weekday=''
        try:
            shopinfor=shop_infor(item[8])
        except:
            failed=open('failed.txt','a')
            failed.write(str(item)+'\n')
            failed.close()
            continue
        f=open('shops.txt','a')
        f.write(str(item+[weekday]+shopinfor)+'\n')
        f.close()
        print(item[0],'ok')
    '''
    users={}
    for line in open('./shops.txt','r'):
        item=eval(line)
        usrid=item[1].split('/')[-1]
        try:
            length=users[usrid]
        except:
            result=get_fans(usrid)
            length=result[0]
            users[usrid]=length
            f=open('fans.txt','a')
            f.write(str([item[0]]+result[1])+'\n')
            f.close()
        f=open('result.txt','a')
        f.write(str(item+[length])+'\n')
        f.close()
        print(item[0],'ok')

def write_to_excel():
    excel=openpyxl.Workbook(write_only=True)
    sheet=excel.create_sheet()
    for line in open('result.txt','r'):
        item=eval(line)
        try:
            item[10]=int(item[10])/10
        except:
            pass
        try:
            item[-6]=int(item[-6])/10
        except:
            pass
        sheet.append(item)
    sheet=excel.create_sheet()
    for line in open('fans.txt','r'):
        item=eval(line)
        sheet.append(item)
    excel.save('result.xlsx')

write_to_excel()
