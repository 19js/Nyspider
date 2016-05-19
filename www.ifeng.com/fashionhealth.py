import requests
import openpyxl
from bs4 import BeautifulSoup
import os

headers = {
        'User-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:39.0) Gecko/20100101 Firefox/39.0',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.5',
        'Accept-Encoding': 'gzip, deflate',
        'Connection': 'keep-alive'}

def topNews(html):
    table=BeautifulSoup(html,'lxml').find('div',attrs={'class':'col_right'}).find_all('div',{'class':'box_01_a clearfix'})
    result=[]
    for item in table:
        title=item.find('h1').get_text()
        url=item.find('a').get('href')
        if '_0' not in url:
            continue
        images=''
        content=''
        page=0
        while True:
            html=requests.get(url.replace('_0','_{}'.format(page)),headers=headers,timeout=30).text.encode('ISO-8859-1','ignore').decode('utf-8','ignore')
            try:
                html=requests.get(url.replace('_0','_{}'.format(page)),headers=headers,timeout=30).text.encode('ISO-8859-1','ignore').decode('utf-8','ignore')
            except:
                break
            page+=1
            soup=BeautifulSoup(html,'html.parser').find('div',id='main_content')
            if soup==None:
                try:
                    soup=BeautifulSoup(html,'lxml').find('div',id='imgBox')
                except:
                    break
                if soup==None:
                    break
            imgs=soup.find_all('img')
            for img in imgs:
                try:
                    images+=img.get('src')+';'
                except:
                    continue
            content+=soup.get_text()
        if content=='':
            continue
        article=[title,images,content]
        result.append(article)
    return result


def recommended(html):
    table=BeautifulSoup(html,'lxml').find('div',attrs={'class':'box_wm'}).find_all('div',{'class':'box_02 clearfix'})
    result=[]
    for item in table:
        title=item.find('h3').get_text()
        url=item.find('a').get('href')
        if '_0' not in url:
            continue
        images=''
        content=''
        page=0
        while True:
            try:
                html=requests.get(url.replace('_0','_{}'.format(page)),headers=headers,timeout=30).text.encode('ISO-8859-1','ignore').decode('utf-8','ignore')
            except:
                break
            page+=1
            soup=BeautifulSoup(html,'html.parser').find('div',id='main_content')
            if soup==None:
                break
            imgs=soup.find_all('img')
            for img in imgs:
                images+=img.get('src')+';'
            content+=soup.get_text()
        if content=='':
            continue
        article=[title,images,content]
        result.append(article)
    return result

def News(html):
    soup=BeautifulSoup(html,'lxml').find('div',{'class':'box_08'}).find('div',id='ba1')
    table=soup.find_all('li')
    try:
        table.append(soup.find('h3'))
    except:
        pass
    result=[]
    for item in table:
        title=item.find('a').get_text()
        try:
            url=item.find('a').get('href')
        except:
            continue
        if '_0' not in url:
            continue
        images=''
        content=''
        page=0
        while True:
            try:
                html=requests.get(url.replace('_0','_{}'.format(page)),headers=headers,timeout=30).text.encode('ISO-8859-1','ignore').decode('utf-8','ignore')
            except:
                break
            page+=1
            soup=BeautifulSoup(html,'html.parser').find('div',id='main_content')
            if soup==None:
                break
            imgs=soup.find_all('img')
            for img in imgs:
                images+=img.get('src')+';'
            content+=soup.get_text()
        if content=='':
            continue
        article=[title,images,content]
        result.append(article)
    return result


def main():
    try:
        os.mkdir('result')
    except:
        pass
    html=requests.get('http://fashion.ifeng.com/health/',headers=headers).text.encode('ISO-8859-1','ignore').decode('utf-8','ignore')
    topnews=topNews(html)
    excel=openpyxl.Workbook(write_only=True)
    sheet=excel.create_sheet()
    for item in topnews:
        sheet.append(item)
    excel.save('result/topnews.xls')
    print('topNews ok')
    recommend=recommended(html)
    excel=openpyxl.Workbook(write_only=True)
    sheet=excel.create_sheet()
    for item in recommend:
        sheet.append(item)
    excel.save('result/recommend.xls')
    print('recommended ok')
    news=News(html)
    excel=openpyxl.Workbook(write_only=True)
    sheet=excel.create_sheet()
    for item in news:
        sheet.append(item)
    excel.save('result/news.xls')
    print('News ok')

main()
