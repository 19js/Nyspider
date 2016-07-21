import requests
from bs4 import BeautifulSoup
import json
import time
import openpyxl

def citys():
    html=open('index.html','r').read()
    table=BeautifulSoup(html,'lxml').find('ul',{'class':'city_info'}).find_all('li')
    f=open('citys.txt','w')
    for li in table:
        for item in li.find_all('a'):
            f.write(item.get_text()+'\n')
    f.close()

def get_store(city):
    result=[]
    page=1
    while True:
        data={
        'cname':city,
        'pid':"",
        'pageIndex':page,
        'pageSize':"100"
        }
        html=requests.post('http://www.kfc.com.cn/kfccda/ashx/GetStoreList.ashx?op=cname',data=data).text
        stores=json.loads(html)['Table1']
        if stores==[]:
            break
        page+=1
        for item in stores:
            result.append(item['storeName']+'|'+item['cityName']+'|'+item['addressDetail']+'|'+item['pro'])
        time.sleep(1)
    return result


def main():
    f=open('result.txt','a')
    for line in open('citys.txt','r'):
        city=line.replace('\n','')
        try:
            result=get_store(city)
        except:
            failed=open('failed.txt','a')
            failed.write(city+'\n')
            failed.close()
            continue
        for item in result:
            f.write(item+'\n')
        print(city,'ok')
    f.close()

def write_to_excel():
    result={}
    excel=openpyxl.Workbook(write_only=True)
    sheet1=excel.create_sheet('1')
    for line in open('result.txt','r'):
        line=line.replace('\n','')
        lists=line.split('|')
        lists[0]=lists[0]+'餐厅'
        try:
            result[lists[1]]+=1
        except:
            result[lists[1]]=1
        sheet1.append(lists)
    sheet2=excel.create_sheet('2')
    for key in result:
        sheet2.append([key,result[key]])
    excel.save('result.xlsx')

write_to_excel()
