import requests
from bs4 import BeautifulSoup
import json
import time
import openpyxl
import jieba
import re

headers = {
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Encoding": "gzip, deflate",
    "Accept-Language": "en-US,en;q=0.5",
    "Connection": "keep-alive",
    "User-Agent": "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:39.0) Gecko/20100101 Firefox/39.0"}

def search(key):
    result=[]
    offset=0
    while offset<100:
        data=requests.get('https://www.zhihu.com/r/search?q={}&type=content&offset={}'.format(key,offset),headers=headers).text
        items=parser(data)
        for item in items:
            question=question_infor(item)
            result.append(question)
            time.sleep(1)
        offset+=10
        print(offset)
    return result

def parser(data):
    htmls=json.loads(data)['htmls']
    result=[]
    for html in htmls:
        item={}
        try:
            soup=BeautifulSoup(html,'html.parser')
            item['title']=soup.find('a').get_text()
            soup=soup.find('div',{'class':'content'})
            item['question-id']=soup.find('meta',{'itemprop':'question-id'}).get('content')
            item['question-url-token']=soup.find('meta',{'itemprop':'question-url-token'}).get('content')
            answer=soup.find('div',{'class':['entry','answer']})
            item['answer-id']=answer.find('meta',{'itemprop':'answer-id'}).get('content')
            item['answer-url-token']=answer.find('meta',{'itemprop':'answer-url-token'}).get('content')
            item['votecount']=answer.find('a',{'class':'zm-item-vote-count'}).get_text()
            try:
                item['author']=answer.find('a',{'class':'author'}).get_text()
            except:
                item['author']='匿名用户'
            item['answer-content']=answer.find('script',{'class':'content'}).get_text()
            item['answer-content']=BeautifulSoup(item['answer-content'],'html.parser').get_text()
            item['date']=answer.find('a',{'class':'time text-muted'}).get('data-tooltip')
            item['answer-comment-count']=answer.find('a',{'class':'js-toggleCommentBox'}).find('span').get_text()
            if '添加评论' in item['answer-comment-count']:
                item['answer-comment-count']=0
            item['copyright']=answer.find('a',{'class':'js-copyright'}).get_text()
            result.append(item)
        except:
            continue
    return result

def question_infor(item):
    html=requests.get('https://www.zhihu.com/question/%s'%(item['question-url-token']),headers=headers).text
    soup=BeautifulSoup(html,'html.parser').find('div',{'class':'zu-main-content-inner'})
    try:
        item['question-content']=soup.find('div',{'class':'zm-editable-content'}).get_text()
    except:
        item['question-content']='-'
    try:
        item['comment-count']=soup.find('div',id='zh-question-meta-wrap').find('a',{'name':'addcomment'}).get_text()
        if '添加评论' in item['comment-count']:
            item['comment-count']=0
    except:
        item['comment-count']='-'
    html=requests.get('https://www.zhihu.com/node/QuestionCommentListV2?params={"question_id":%s}'%(item['question-id']),headers=headers).text
    table=BeautifulSoup(html,'html.parser').find_all('div',{'class':'zm-item-comment'})
    item['comments']=[]
    for div in table:
        comment={}
        comment['comment-id']=div.get('data-id')
        comment['comment-content']=div.find('div',{'class':'zm-comment-content'}).get_text()
        item['comments'].append(comment)
    return item

def get_comments(answerid):
    comments=[]
    page=1
    pre=[]
    while True:
        try:
            html=requests.get('https://www.zhihu.com/r/answers/%s/comments?page=%s'%(answerid,page),headers=headers).text
            data=json.loads(html)['data']
            if data==pre:
                break
            pre=data
            for item in data:
                comments.append(item['content'])
            print('Get comments',answerid,page,'ok')
            page+=1
            time.sleep(1)
        except:
            break
    return comments

def write_to_excel(items):
    excel=openpyxl.Workbook(write_only=True)
    sheet1=excel.create_sheet('question')
    sheet2=excel.create_sheet("answer")
    question_keys=['title', 'question-id','question-url-token','question-content', 'comment-count']
    answer_keys=['question-id','answer-id','answer-url-token','author', 'votecount', 'answer-content','answer-comment-count', 'copyright', 'date']
    sheet1.append(question_keys+['comment-id','comment-content'])
    sheet2.append(answer_keys)
    for item in items:
        question=[]
        for key in question_keys:
            question.append(item[key])
        if len(item['comments'])==0:
            sheet1.append(question)
        for comment in item['comments']:
            sheet1.append(question+[comment['comment-id'],comment['comment-content']])
        answer=[]
        for key in answer_keys:
            answer.append(item[key])
        sheet2.append(answer)
    excel.save('result.xlsx')

result=search("zhihu")
write_to_excel(result)
