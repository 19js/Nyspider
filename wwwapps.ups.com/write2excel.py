import openpyxl

def load_data():
    keys=[line.replace('\n','').replace(' ','') for line in open('data','r')]
    data={}
    for line in open('result.txt','r'):
        line=line.replace('\n','').split('-')
        try:
            data[line[0]][line[1]]=int(line[-1])
        except:
            data[line[0]]={}
            data[line[0]][line[1]]=int(line[-1])
        try:
            data[line[1]][line[0]]=int(line[-1])
        except:
            data[line[1]]={}
            data[line[1]][line[0]]=int(line[-1])
    return keys,data

def write_to_excel():
    keys,data=load_data()
    excel=openpyxl.Workbook(write_only=True)
    sheet=excel.create_sheet()
    line=['']
    for key in keys:
        if len(key)==4:
            key='0'+key
        line.append(key)
    sheet.append(line)
    for key in keys:
        if len(key)==4:
            key='0'+key
        line=[key]
        for another_key in keys:
            if len(another_key)==4:
                another_key='0'+another_key
            if key==another_key:
                line.append(1)
            else:
                try:
                    line.append(data[key][another_key])
                except:
                    line.append('')
        sheet.append(line)
    sheet=excel.create_sheet()
    line=['']
    for key in keys:
        if len(key)==4:
            key='0'+key
        line.append(key)
    sheet.append(line)
    for key in keys:
        if len(key)==4:
            key='0'+key
        line=[key]
        for another_key in keys:
            if len(another_key)==4:
                another_key='0'+another_key
            if key==another_key:
                line.append(1)
            else:
                try:
                    value=data[key][another_key]
                    if value!=1:
                        value=0
                    line.append(value)
                except:
                    line.append('')
        sheet.append(line)
    excel.save('result.xlsx')

write_to_excel()
