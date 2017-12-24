import requests
import time
import json
import re
import openpyxl
import os
from selenium import webdriver
from bs4 import BeautifulSoup

def get_product_comments(itemId, sellerId, product):
    url = 'https://rate.tmall.com/list_detail_rate.htm?itemId={}&sellerId={}&order=1&currentPage={}&append=0&content=1&tagId=&posi=&picture='
    page = 1
    rateCount = ''
    while True:
        try:
            res_text = requests.get(url.format(itemId, sellerId, page)).text
            res_text = res_text.replace('"rateDetail":', '')
            data = json.loads(res_text)
            if rateCount == '':
                rateCount = data['rateCount']
            rateList = data['rateList']
        except Exception as e:
            print(e)
            continue
        with open('result/%s.txt' % (itemId), 'a') as f:
            for item in rateList:
                item['rateCount'] = rateCount
                item['itemId'] = itemId
                item['product'] = product
                rateDate = item['rateDate']
                if '2017-12' not in rateDate:
                    break
                f.write(str(item) + '\n')
        if '2017-12' not in rateDate:
            break
        if page == data['paginator']['lastPage']:
            break
        print(itemId, sellerId, page, len(rateList), 'OK')
        if(page == 99):
            break
        page += 1
        time.sleep(1)


def crawl_comments():
    for line in open('./products.txt', 'r'):
        line = eval(line)
        get_product_comments(line[2], line[1].split('=')[-1], line)

def write_comments_to_excel():
    excel = openpyxl.Workbook()
    sheet = excel.create_sheet()
    for line in load_txt():
        # line=ILLEGAL_CHARACTERS_RE.sub(r'', line)
        # line=eval(line)
        try:
            sheet.append(line)
        except Exception as e:
            print(e)
    excel.save('评论.xlsx')

def write_product_to_excel():
    excel = openpyxl.Workbook()
    sheet = excel.create_sheet()
    for line in open('result.txt','r'):
        #line=ILLEGAL_CHARACTERS_RE.sub(r'', line)
        try:
            line=eval(line)
        except Exception as e:
            line=line.replace('[','').replace(']','').replace("'",'').split(',')
            
        try:
            sheet.append(line)
        except Exception as e:
            print(e)
    excel.save('产品信息.xlsx')


def load_txt():
    need_keys = ['itemId','id','auctionSku','displayUserNick','userVipLevel','rateDate','rateContent']
    exists = {}
    for ID in [eval(line)[2] for line in open('./products.1.txt','r')]:
        filename='result/%s.txt'%ID
        try:
            comment_f=open(filename,'r')
        except:
            print(ID,'FAILED')
            continue
        for line in comment_f:
            line = eval(line)
            product = line['product']
            product_id = product[2]
            if product_id not in exists:
                comment_item = product
                pic_num=line['rateCount']['picNum']
                append_num=line['rateCount']['used']
                product+=[pic_num,append_num]
                with open('products.txt', 'a') as f:
                    f.write(str(comment_item) + '\n')
                exists[product_id] = 1
            item = []
            for key in need_keys:
                try:
                    item.append(line[key])
                except:
                    item.append('')
            try:
                image_num = len(line['pics'])
            except:
                image_num = 0
            item.append(image_num)
            yield [product[4],product[2]]+item
            
def get_product_info():
    browser=webdriver.Chrome('../chromedriver')
    browser.get('https://tmall.com')
    input("OK?")
    for line in open('./products.txt','r'):
        item=eval(line)
        url='https:'+item[5]
        browser.get(url)
        input('OK?')
        html=browser.page_source
        soup=BeautifulSoup(html,'lxml')
        try:
            J_CollectCount=soup.find('span',id='J_CollectCount').get_text()
        except:
            J_CollectCount='-'
        try:
            J_PostageToggleCont=soup.find('div',id='J_PostageToggleCont').get_text()
        except:
            J_PostageToggleCont='-'
        try:
            tm_count=soup.find('li',{'class':'tm-ind-emPointCount'}).find('span',{'class':'tm-count'}).get_text()
        except:
            tm_count='-'
        print([J_CollectCount,J_PostageToggleCont,tm_count])
        item+=[J_CollectCount,J_PostageToggleCont,tm_count]
        f=open('result.txt','a')
        f.write(str(item)+'\n')
        f.close()


if __name__ == '__main__':
    write_product_to_excel()
