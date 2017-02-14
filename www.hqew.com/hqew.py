import requests
from bs4 import BeautifulSoup
import threading
import time
import openpyxl
import random

user_agents=['Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.87 Safari/537.36'
             ,'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:39.0) Gecko/20100101 Firefox/39.0'
             ,'Mozilla/5.0 (X11; Ubuntu; Linux i686 on x86_64; rv:41.0) Gecko/20100101 Firefox/41.0'
             ,'Mozilla/5.0 (compatible; MSIE 8.0; Windows NT 6.3; Win64; x64)']

def get_headers():
    headers = {
        "X-Forwarded-For":'%s.%s.%s.%s'%(random.randint(0, 255),random.randint(0, 255),random.randint(0, 255),random.randint(0, 255)),
        'Host':"www.hqew.com",
        'User-Agent':random.choice(user_agents),
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.5',
        'Accept-Encoding': 'gzip, deflate'}
    return headers

def load_excel():
    excel=openpyxl.load_workbook('./data.xlsx')
    names=excel.get_sheet_names()
    sheet=excel.get_sheet_by_name(names[0])
    lines=[]
    try:
        count=input('输入抓取线程数:')
        count=int(count)
    except:
        count=10
    for row in sheet.rows:
        line=[str(row[0].value).replace('\r','').replace('\n',''),str(row[1].value).replace('\r','').replace('\n',''),str(row[2].value)]
        lines.append(line)
        if len(lines)<count:
            continue
        yield lines
        lines.clear()
    yield lines

class Parser(threading.Thread):
    def __init__(self,line):
        super(Parser,self).__init__()
        self.line=line

    def run(self):
        try:
            result=parser(self.line[2],self.line[1])
        except:
            result=[0,0]
        if '采集公司' in str(self.line):
            result=['库存','标签']
        self.line=self.line+result

def parser(company,url):
    try:
        html=requests.get(url,headers=get_headers(),timeout=20).text
    except:
        return [0,0]
    table=BeautifulSoup(html,'lxml').find('div',{'class':'list'}).find_all('ul')
    result=[0,'']
    for ul in table:
        try:
            name=ul.find('a',{'class':'company'}).get_text().replace('\r','').replace('\n','').replace('\t','').replace(' ','')
            num=ul.find('li',{'class':'col10'}).get_text().replace('\r','').replace('\n','').replace('\t','').replace(' ','')
            num=int(num)
            label=ul.find('li',{'class':'col4'})
        except:
            continue
        if label is None:
            label=''
        elif '供应商承诺该库存为现货库存' in str(label):
            label='现货'
        elif '供应商承诺该库存为原装现货库存' in str(label):
            label='原装'
        else:
            label=''
        if name in company:
            result[0]+=num
            if label not in result[1]:
                result[1]+=label
    if result[1]=='':
        result[1]=0
    return result

def write_to_excel():
    excel=openpyxl.Workbook(write_only=True)
    sheet=excel.create_sheet()
    for line in open('result.txt','r',encoding='utf-8'):
        try:
            item=eval(line)
            sheet.append(item)
        except:
            continue
    excel.save('result.xlsx')

def main():
    f=open('result.txt','w',encoding='utf-8')
    count=0
    for lines in load_excel():
        threads=[]
        for line in lines:
            work=Parser(line)
            work.setDaemon(True)
            threads.append(work)
        for work in threads:
            work.start()
        for work in threads:
            work.join()
        for work in threads:
            f.write(str(work.line)+'\r\n')
            count+=1
        print(count,'ok')
    f.close()

main()
write_to_excel()
print('完成')
time.sleep(200)
