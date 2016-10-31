import requests
from bs4 import BeautifulSoup
import openpyxl
import threading
import time
import random
from openpyxl import load_workbook


thread_num=3 #线程数量
switch_time=5 #ip切换时间间隔
sleep_time=2 #线程时间

def get_headers():
    user_agents=['Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:44.0) Gecko/20100101 Firefox/44.0','Mozilla/5.0 (Windows NT 6.1; WOW64)','AppleWebKit/537.11 (KHTML, like Gecko)','Chrome/23.0.1271.97','Safari/537.11']
    headers = {
        'Host':"tur.bizdirlib.com",
        'Accept':"text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Encoding": "gzip, deflate",
        "Accept-Language": "en-US,en;q=0.5",
        "User-Agent":user_agents[random.randint(0,len(user_agents))]}
    return headers

def get_proxies():
    proxyHost = "proxy.abuyun.com"
    proxyPort = "9010"
    # 代理隧道验证信息
    proxyUser = "H55D5746G8C8151P"
    proxyPass = "145C18388AF8DBAC"

    proxyMeta = "http://%(user)s:%(pass)s@%(host)s:%(port)s" % {
      "host" : proxyHost,
      "port" : proxyPort,
      "user" : proxyUser,
      "pass" : proxyPass,
    }
    proxies = {
        "http"  : proxyMeta,
        "https" : proxyMeta,
    }
    return proxies

class Crawler(threading.Thread):
    def __init__(self,url):
        super(Crawler,self).__init__()
        self.url=url

    def run(self):
        self.status=True
        try:
            self.result=self.parser()
        except:
            self.status=False

    def parser(self):
        html=requests.get(self.url,headers=get_headers(),timeout=20,proxies=get_proxies()).text
        table=BeautifulSoup(html,'html.parser').find('div',{'class':'content'}).find('ul').find_all('li')
        result={}
        result['url']=self.url
        for li in table:
            try:
                item=li.get_text().replace('\r','').replace('\n','')
                tag=item.split(':')[0]
                value=item.split(':')[-1]
                result[tag]=value
            except:
                continue
        return result

def load_urls(urls_filename):
    '''
    workbook=load_workbook(filename='urls.xlsx')
    names=workbook.get_sheet_names()
    sheet=workbook.get_sheet_by_name(names[0])
    urls=[]
    for row in sheet.rows[1:]:
        urls.append(row[0].value)
    return urls
    '''
    urls=[]
    for line in open(urls_filename,'r',encoding='utf-8'):
        line=line.replace('\r','').replace('\n','')
        urls.append(line)
    return urls

def switch_ip():
    html=requests.get('http://proxy.abuyun.com/switch-ip',proxies=get_proxies()).text
    print("Switch ip ",html)
    global t
    global switch_time
    t=threading.Timer(switch_time,switch_ip)
    t.start()

def write_to_excel(result):
    excel=openpyxl.Workbook(write_only=True)
    sheet=excel.create_sheet()
    keys=['url','Company Name','Country','Address','International Area Code','Phone','Fax','Contact','Email','Website','Area']
    sheet.append(keys)
    for item in result:
        line=[]
        for key in keys:
            try:
                line.append(item[key])
            except:
                line.append('')
        sheet.append(line)
    timenow=time.strftime('%Y%m%d_%H%M%S')
    excel.save(timenow+'_result.xlsx')

t=threading.Timer(switch_time,switch_ip)
t.setDaemon(True)
t.start()

def main():
    urls_filename='urls.txt'
    urls=load_urls(urls_filename)
    result=[]
    failed_count={}
    while len(urls):
        threadings=[]
        for i in range(thread_num):
            try:
                url=urls.pop()
            except:
                break
            crawler=Crawler(url)
            crawler.setDaemon(True)
            threadings.append(crawler)
        for crawler in threadings:
            crawler.start()
        for crawler in threadings:
            crawler.join()
        for crawler in threadings:
            url=crawler.url
            if crawler.status==False:
                try:
                    count=failed_count[url]
                except:
                    failed_count[url]=1
                    urls.append(url)
                    continue
                if count<4:
                    failed_count[url]+=1
                    urls.append(url)
                else:
                    failed=open('failed.txt','a',encoding='utf-8')
                    failed.write(crawler.url+'\n')
                    failed.close()
                continue
            print(url,'ok')
            result.append(crawler.result)
        time.sleep(sleep_time)
    write_to_excel(result)
    print('完成')

main()
