import requests
from bs4 import BeautifulSoup
import openpyxl

def citys():
    f=open('citys.txt','a')
    for pid in range(6,33):
        html=requests.get('http://www.dicos.com.cn/index.php?c=page&m=getcityhtml&iscity=1&pid=%s'%pid).text
        table=BeautifulSoup(html,'lxml').find_all('option')
        for item in table:
            f.write(item.get_text()+'|'+item.get('value')+'\n')
    f.close()

def get_store(citycode):
    html=requests.get('http://www.dicos.com.cn/index.php?c=page&m=getstorehtml&waimai=0&mProvince=3&mCity=%s'%citycode).text
    table=BeautifulSoup(html,'lxml').find_all('tr')
    result=[]
    for item in table:
        text=''
        for td in item.find_all('td')[1:4]:
            text+='|'+td.get_text()
        result.append(text.replace('\r','').replace('\n',''))
    return result

def main():
    f=open('result.txt','a')
    for line in open('citys.txt'):
        line=line.replace('\n','')
        try:
            result=get_store(line.split('|')[-1])
        except:
            failed=open('failed.txt','a')
            failed.write(line+'\n')
            failed.close()
            continue
        for item in result:
            f.write(line+item+'\n')
        print(line,'ok')
    f.close()

def write_to_excel():
    result={}
    excel=openpyxl.Workbook(write_only=True)
    sheet1=excel.create_sheet('1')
    for line in open('result.txt','r'):
        line=line.replace('\n','')
        lists=line.split('|')
        try:
            result[lists[0]]+=1
        except:
            result[lists[0]]=1
        sheet1.append(lists)
    sheet2=excel.create_sheet('2')
    for key in result:
        sheet2.append([key,result[key]])
    excel.save('result.xlsx')

write_to_excel()
