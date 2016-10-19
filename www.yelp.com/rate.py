import requests
from bs4 import BeautifulSoup
import openpyxl
import time

headers = {
    'User-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:39.0) Gecko/20100101 Firefox/39.0',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.5',
    'Accept-Encoding': 'gzip, deflate',
    'Connection': 'keep-alive'}


def get_infor(url):
    html=requests.get(url,headers=headers).text
    soup=BeautifulSoup(html,'lxml').find('div',{'class':'main-content-wrap--full'})
    baseinfor=soup.find('div',{'class':'top-shelf'})
    name=baseinfor.find('h1').get_text()
    Claimed=baseinfor.find('span',{'class':'claim-status_teaser'}).get_text()
    if 'Unclaimed' in Claimed:
        Claimed='0'
    else:
        Claimed='1'
    try:
        rating=baseinfor.find('div',{'class':'rating-very-large'}).find('i').get('title').replace('star rating','')
    except:
        rating=''
    try:
        review=baseinfor.find('span',{'class':'review-count'}).find('span').get_text()
    except:
        review=''
    try:
        price_range=baseinfor.find('div',{'class':'price-category'}).find('span',{'class':'price-range'}).get_text().replace(' ','')
        price_range=len(price_range)
    except:
        price_range=''
    try:
        address=baseinfor.find('div',{'class':'map-box-address'}).find('address').get_text()
    except:
        address=''
    ylist=soup.find('div',{'class':'bordered-rail'}).find('div',{'class':'short-def-list'}).find_all('dl')
    Waiter_Services=''
    Drive_Thru=''
    for dl in ylist:
        if 'Waiter Service' in str(dl):
            if 'Yes' in str(dl):
                Waiter_Services='1'
            else:
                Waiter_Services='0'
        if 'Drive-Thru' in str(dl):
            if 'Yes' in str(dl):
                Drive_Thru='1'
            else:
                Drive_Thru='0'
    ratelist=eval(soup.find('div',id='rating-details-modal-content').get('data-monthly-ratings'))
    rate_result={''}
    r_2012=['']*12
    r_2013=['']*12
    r_2014=['']*12
    r_2015=['']*12
    r_2016=['']*12
    for key in ratelist:
        if key=='2012':
            for item in ratelist['2012']:
                r_2012


def load_excel(filename):
    excel=openpyxl.load_workbook(filename)
    names=excel.get_sheet_names()
    sheet=excel.get_sheet_by_name(names[0])
    items=[]
    for row in sheet.rows[1:]:
        business_id=row[0].value
        name=row[1].value
        url=row[2].value
        items.append([business_id,name,url])
    return items

def write_to_excel(result):
    excel=openpyxl.Workbook(write_only=True)
    sheet=excel.create_sheet()
    for item in result:
        sheet.append(item)
    excel.save('result.xlsx')

def main():
    items=load_excel('Food2.1.xlsx')
    result=[]
    for item in items:
        try:
            line=get_infor(item[-1])
        except:
            failed=open('failed.txt','a',encoding='utf-8')
            failed.write(str(item)+'\n')
            failed.close()
            continue
        result.append(line)
        print(item,'ok')
    write_to_excel(result)

get_infor('https://www.yelp.com/biz/arbys-canoga-park')
