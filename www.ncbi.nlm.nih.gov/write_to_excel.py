import openpyxl

def write_to_excel():
    excel=openpyxl.Workbook(write_only=True)
    sheet=excel.create_sheet()
    count=0
    filecount=1
    exist=[]
    for line in open('result.txt','r'):
        line=line.replace('\r\n','')
        item=eval(line)
        if item[0] in exist:
            continue
        exist.append(item[0])
        sheet.append(item)
        count+=1
        print(count)
        if count%100000==0:
            excel.save('%s.xlsx'%filecount)
            filecount+=1
            excel=openpyxl.Workbook(write_only=True)
            sheet=excel.create_sheet()
    excel.save('%s.xlsx'%filecount)

write_to_excel()
