import requests
import json
import openpyxl
import time
from bs4 import BeautifulSoup

def login():
    data={
    'user_name':'',
    'password':'',
    'net_auto_login':1,
    '_post_type':'ajax',
    'return_url':'https://www.jisilu.cn'
    }
    session=requests.session()
    session.post('https://www.jisilu.cn/account/ajax/login_process/',data=data).text
    return session

def get_A(session):
    data={
    'is_funda_search':"0",
    'fundavolume':"100",
    'maturity':"",
    'amarket':["sh",'sz'],
    'coupon_descr':["+3.0%",'+3.2%','+3.5%','+4.0%','other'],
    'rp':"50",
    'page':'1'
    }
    timestr=str(time.time()).replace('.','')
    html=session.post('https://www.jisilu.cn/data/sfnew/funda_list/?___t=%s'%(timestr),data=data).text
    data=json.loads(html)['rows']
    result=[]
    header=['代码', '名称', '现价', '涨幅', '成交额(万元)', '净值', '折价率', '利率规则', '本期利率', '下期利率', '修正收益率', '剩余年限', '参考指数', '指数涨幅', '下折母基需跌', '理论下折收益', '上折母基需涨', '整体溢价率', 'T-1溢价率', 'T-2溢价率', 'A份额(万份)', 'A新增(万份)', 'A:B', '下次定折']
    result.append(header)
    keys=['funda_id','funda_name','funda_current_price','funda_increase_rt','funda_volume','funda_value'
        ,'funda_discount_rt','coupon_descr_s','funda_coupon','funda_coupon_next','funda_profit_rt_next','funda_left_year'
        ,'funda_index_name','funda_index_increase_rt','funda_lower_recalc_rt','lower_recalc_profit_rt'
        ,'fundb_upper_recalc_rt','funda_base_est_dis_rt','funda_base_est_dis_rt_t1','funda_base_est_dis_rt_t2'
        ,'funda_amount','funda_amount_increase','abrate','next_recalc_dt']
    for item in data:
        item=item['cell']
        line=[]
        for key in keys:
            if key=='next_recalc_dt':
                try:
                    value=BeautifulSoup(item[key],'html.parser').get_text()
                except:
                    value=item[key]
                line.append(value)
                continue
            try:
                line.append(item[key])
            except:
                line.append('')
        result.append(line)
    return result

def get_B(session):
    data={
    'rp':"50",
    'page':'1'
    }
    timestr=str(time.time()).replace('.','')
    html=session.post('https://www.jisilu.cn/data/sfnew/fundb_list/?___t=%s'%(timestr),data=data).text
    data=json.loads(html)['rows']
    result=[]
    header=['代码', '名称', '现价', '涨幅', '成交额(万元)', '估值', '净值', '溢价率', '剩余年限', '利率规则', '价格杠杆', '净值杠杆', '融资成本', '参考指数', '指数涨幅', '下折母基需跌', '上折母基需涨', '整体溢价率', 'A:B', '母基净值']
    result.append(header)
    keys=['fundb_id','fundb_name','fundb_current_price','fundb_increase_rt','fundb_volume','b_est_val'
        ,'fundb_value','fundb_discount_rt','fundb_left_year','coupon_descr_s','fundb_price_leverage_rt'
        ,'fundb_net_leverage_rt','fundb_capital_rasising_rt','fundb_index_name','fundb_index_increase_rt'
        ,'fundb_lower_recalc_rt','fundb_upper_recalc_rt','fundb_base_est_dis_rt','abrate','fundm_value']
    for item in data:
        item=item['cell']
        line=[]
        for key in keys:
            try:
                line.append(item[key])
            except:
                line.append('')
        result.append(line)
    return result

def get_M(session):
    data={
    'rp':"50",
    'page':'1'
    }
    timestr=str(time.time()).replace('.','')
    html=session.post('https://www.jisilu.cn/data/sfnew/fundm_list/?___t=%s'%(timestr),data=data).text
    data=json.loads(html)['rows']
    result=[]
    header=['母基代码', '母基名称', '母基净值', '净值日期', '跟踪指数', '上折', '下折', '下折需跌', '创立日期', '到期日', 'A基代码', 'A基名称', '本期利率', '下期利率', '利率规则', 'B基代码', 'B基名称', 'A:B', '整体溢价率', '基金管理人']
    result.append(header)
    keys=['base_fund_id','base_fund_nm','price','last_chg_dt','index_nm','upper_recalc_price'
        ,'lower_recalc_price','base_lower_recalc_rt','issue_dt','maturity_dt','fundA_id','fundA_nm'
        ,'coupon','coupon_next','coupon_descr_s','fundB_id'
        ,'fundB_nm','abrate','base_est_dis_rt','fund_company_nm']
    for item in data:
        item=item['cell']
        line=[]
        for key in keys:
            try:
                line.append(item[key])
            except:
                line.append('')
        result.append(line)
    return result

def get_F(session):
    data={
    'is_search':"0",
    'avolume':"100",
    'bvolume':"100",
    'market':["sh","sz"],
    'ptype':"price",
    'rp':"50",
    'page':"1"
    }
    timestr=str(time.time()).replace('.','')
    html=session.post('https://www.jisilu.cn/data/sfnew/arbitrage_vip_list/?___t=%s'%timestr,data=data).text
    data=json.loads(html)['rows']
    result=[]
    header=['A代码', 'A名称', 'A价格', 'A涨幅', 'A成交(万元)', 'A新增(万份)', 'B代码', 'B名称', 'B价格', 'B涨幅', 'B成交(万元)', 'B新增(万份)', 'A:B', '合并价格', '合并溢价', '母基代码', '母基名称', '母基净值', '估算净值', '跟踪指数', '指数涨幅', '估值仓位', '前日仓位', '申购费', '赎回费']
    result.append(header)
    keys=['fundA_id','fundA_nm','sell1A','increase_rtA','fundA_volume','fundA_amount_increase',
        'fundB_id','fundB_nm','sell1B','increase_rtB','fundB_volume','fundB_amount_increase',
        'abrate','merge_price','est_dis_rt','base_fund_id','base_fund_nm','base_nav','base_est_val',
        'index_nm','idx_incr_rt','asset_ratio','asset_ratio_last','apply_fee','redeem_fee']
    for item in data:
        item=item['cell']
        line=[]
        for key in keys:
            try:
                line.append(item[key])
            except:
                line.append('')
        result.append(line)
    return result

def get_ETF(session):
    data={
    'rp':"50",
    'page':'1'
    }
    timestr=str(time.time()).replace('.','')
    html=session.post('https://www.jisilu.cn/jisiludata/etf.php?___t=%s'%(timestr),data=data).text
    data=json.loads(html)['rows']
    result=[]
    header=['代码', '名称', '现价', '涨幅', '成交额(万元)', '指数', '指数PE', '指数PB', '指数涨幅', '估值', '净值', '净值日期', '溢价率', '最小申赎单位(万份)', '份额(万份)', '规模变化(亿元)', '规模(亿元)']
    result.append(header)
    keys=['fund_id','fund_nm','price','increase_rt','volume','index_nm'
        ,'pe','pb','index_increase_rt','estimate_value','fund_nav'
        ,'nav_dt','discount_rt','creation_unit','amount'
        ,'unit_incr','unit_total']
    for item in data:
        item=item['cell']
        line=[]
        for key in keys:
            try:
                line.append(item[key])
            except:
                line.append('')
        result.append(line)
    return result

def write_to_excel():
    excel=openpyxl.Workbook(write_only=True)
    session=login()
    A_result=get_A(session)
    sheet=excel.create_sheet('A类')
    for item in A_result:
        sheet.append(item)
    B_result=get_B(session)
    sheet=excel.create_sheet('B类')
    for item in B_result:
        sheet.append(item)
    M_result=get_M(session)
    sheet=excel.create_sheet('母基')
    for item in M_result:
        sheet.append(item)
    F_result=get_F(session)
    sheet=excel.create_sheet('分级套利')
    for item in F_result:
        sheet.append(item)
    ETF_result=get_ETF(session)
    sheet=excel.create_sheet("ETF")
    for item in ETF_result:
        sheet.append(item)
    excel.save('result.xls')

'''定时模块

while True:
    timenow=time.strftime("%Y-%m-%d %H:%M:%S",time.localtime())
    if '21:26:' in timenow:#修改 21:26:,
        count=0
        while True:
            try:
                write_to_excel()
            except:
                count+=1
                if count==5:
                    print('Failed')
                    break
    time.sleep(10)
'''
count=0
while True:
    try:
        write_to_excel()
        print('OK')
        break
    except:
        count+=1
        if count==5:
            print('Failed')
            break

time.sleep(50)
