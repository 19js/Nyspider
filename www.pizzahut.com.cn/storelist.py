import requests
from bs4 import BeautifulSoup
import urllib
import openpyxl

def citys():
    html=open('index.html','r').read()
    table=BeautifulSoup(html,'lxml').find_all('div',{'class':'city_window'})[1].find_all('a')
    f=open('citys.txt','w')
    for item in table:
        f.write(item.get_text()+'\n')
    f.close()

def get_store(city):
    city=urllib.parse.quote(city)
    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.5',
        'Accept-Encoding': 'gzip, deflate',
        'Cookie':"NSC_CX_QfstjtufodzHspvq=ffffffff09320b0745525d5f4f58455e445a4a423660; _u_=1; __RequestVerificationToken=tOMoZty3Jp6D53oSF-NqlfyAlPa0sRNndZ7PNG5iPrWgM_ngcVFEOP79uEvHJGuqlHDoAA3WDd1MN9QA8ZEhpurYLA0WSkuyswlEO9Nj9oqeMWnu84Q1fyQQYx5-vjq-73NNZXJJLcF9jq3fjB_dsw2; iplocation={}%7C0%7C0".format(city),
        'User-Agent':"Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:44.0) Gecko/20100101 Firefox/44.0",
        'Connection': 'keep-alive'}
    page=1
    result=[]
    while True:
        data={
        'pageIndex':page,
        'pageSize':"100",
        'keyword':"输入餐厅地址或餐厅名称"
        }
        html=requests.post('http://www.pizzahut.com.cn/StoreList/Index',headers=headers,data=data).text
        soup=BeautifulSoup(html,'lxml').find_all('li')
        items=[]
        for li in soup:
            item=''
            try:
                for p in li.find('div',{'class':'re_RNew'}).find_all('p'):
                    item+='|'+p.get_text()
            except:
                continue
            items.append(item)
        if items==[]:
            break
        result+=items
        page+=1
    return result

def main():
    f=open('result.txt','a')
    for line in open('citys.txt','r'):
        city=line.replace('\n','')
        try:
            result=get_store(city)
        except:
            failed=open('failed.txt','a')
            failed.write(city+'\n')
            failed.close()
            continue
        for item in result:
            f.write(city+item+'\n')
        print(city,'ok')
    f.close()

def write_to_excel():
    result={}
    excel=openpyxl.Workbook(write_only=True)
    sheet1=excel.create_sheet('1')
    for line in open('result.txt','r'):
        line=line.replace('\n','')
        lists=line.split('|')
        try:
            result[lists[1]]+=1
        except:
            result[lists[1]]=1
        sheet1.append(lists)
    sheet2=excel.create_sheet('2')
    for key in result:
        sheet2.append([key,result[key]])
    excel.save('result.xlsx')

write_to_excel()
